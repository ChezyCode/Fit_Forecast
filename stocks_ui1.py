import streamlit as st
from io import BytesIO
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import yfinance as yf
import logging
import mpmath as mp
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

mp.dps = 100

logging.basicConfig(
    level=logging.DEBUG, 
    format="%(asctime)s | %(levelname)s | %(filename)s:%(lineno)d | %(message)s",
    handlers=[logging.StreamHandler()]
)

# Getting data from Yahoo finance (previously in getting_data.py)
def get_data(stock_name, start_date, end_date):
    data: pd.DataFrame = yf.download(stock_name, start=start_date, end=end_date, auto_adjust=False)
    closing_prices = data['Close']
    s_obs = closing_prices.to_numpy()
    return s_obs

def determine_v_n(Sn,Sn_1):
    v_n = (Sn - Sn_1) / 1 #delta_t = 1

    # Pemeriksaan untuk mencegah pembagian dengan nol 
    if abs(v_n) < 1e-12:
        return 1e-12 
    return v_n

def determine_alpha_n(Sn_minus_4, Sn_minus_3, Sn_minus_2, Sn_minus_1):
    AA = (Sn_minus_2 - 2 * Sn_minus_3 + Sn_minus_4)
    BB = (Sn_minus_1 - Sn_minus_2)
    CC = (Sn_minus_1 - 2 * Sn_minus_2 + Sn_minus_3)
    DD = (Sn_minus_2 - Sn_minus_3)

    alpha_pembilang = (AA * BB) - (CC * DD)
    alpha_penyebut = DD * BB * (DD - BB) 

    # Pemeriksaan untuk mencegah pembagian dengan nol 
    if abs(alpha_penyebut) < 1e-12:
        return 1e-12  
    return (alpha_pembilang/alpha_penyebut)

def determine_beta_n(Sn_minus_3, Sn_minus_2, Sn_minus_1, alpha_n):
    CC = (Sn_minus_1 - 2 * Sn_minus_2 + Sn_minus_3)
    BB = (Sn_minus_1 - Sn_minus_2)

    # Pemeriksaan untuk mencegah pembagian dengan nol 
    if abs(BB) < 1e-12:
        return 1e-12 
    
    return (CC-(alpha_n * (BB**2)))/(BB * 1) #delta_t = 1

def determine_h_n(v_1, alpha_n, beta_n):
    
    # Pemeriksaan untuk mencegah pembagian dengan nol 
    if abs(alpha_n) < 1e-12:
        alpha_n = 1e-12
    if abs(v_1) < 1e-12:
        v_1 = 1e-12
    
    try:
        h_n = abs((v_1 + (beta_n / alpha_n) / v_1))
        return h_n
    except (ZeroDivisionError) as e:
        logging.warning(f"Error in determine_h_n: {e}. Using fallback value.")
        return 1.0

def determine_s_n(s1, alpha, beta, h, condition_1, s_n, v_n, v_1):
    logging.debug(f"determine_s_n called with: s1={s1}, alpha={alpha}, beta={beta}, h={h}, condition_1={condition_1}, s_n={s_n}, v_n={v_n}, v_1={v_1}")

    # Pemeriksaan untuk mencegah pembagian dengan nol 
    if abs(alpha) < 1e-12:
        alpha = 1e-12
    if abs(beta) < 1e-12:
        beta = 1e-12

    condition_2 = v_n > v_1
    condition_3 = s_n > s1
 
    try:
        if condition_1 > 0 and condition_2 and condition_3:
            s_n = s1 - (1/alpha) * mp.log(mp.fabs((mp.exp(beta) - h) / (1 - h)))
        if condition_1 > 0 and condition_2 and not condition_3:
            s_n = s1 + mp.fabs(1/alpha) * (mp.fabs(beta)/beta) * mp.log(mp.fabs((mp.exp(beta) - h) / (1 - h)))
        if condition_1 < 0 and condition_2 and condition_3:
            s_n = s1 - (1/alpha) * mp.log(mp.fabs((mp.exp(beta) + h) / (1 + h)))
        if condition_1 < 0 and condition_2 and not condition_3:
            s_n = s1 - mp.fabs(1/alpha) * (mp.fabs(beta)/beta) * mp.log(mp.fabs((mp.exp(beta) + h) / (1 + h)))
        if condition_1 > 0 and not condition_2 and condition_3:
            s_n = s1 - (1/alpha) * (beta/mp.fabs(beta)) * mp.log(mp.fabs((mp.exp(beta) -h) / (1 - h)))
        if condition_1 > 0 and not condition_2 and not condition_3:
            s_n = s1 - mp.fabs(1/alpha) * mp.log(mp.fabs((mp.exp(-mp.fabs(beta)) - h) / (1 - h)))
        if condition_1 < 0 and not condition_2 and condition_3:
            s_n = s1 + (1/alpha) * (beta/mp.fabs(beta)) * mp.log(mp.fabs(mp.exp(-mp.fabs(beta)) + h) / (1 + h))
        if condition_1 < 0 and not condition_2 and not condition_3:
            s_n = s1 + mp.fabs(1/alpha) * mp.log(mp.fabs(mp.exp(-mp.fabs(beta)) + h) / (1 + h))
    except (ZeroDivisionError) as e:
        logging.error(f'Error in determine_s_n: {e}. Using fallback value.')
        s_n = s1  # menggunakan nilai sebelumnya sebagai nilai fallback

    logging.debug(f'determine_s_n result: s_n={s_n}')
    return s_n

def determine_MAPE_list(actual: list, predicted: list) -> list:
    logging.debug(f'actual: {actual}, len {len(actual)}')
    logging.debug(f"predicted: {predicted}, len {len(predicted)}")
    min_len = min(len(actual), len(predicted))
    actual = actual[:min_len]
    predicted = predicted[:min_len]
    num_of_cases = len(actual)
    sum_of_percentage_error = 0
    mape_list = []
    for i in range(num_of_cases):
        if actual[i] == 0:
            continue  # Skip jika nilai aktual = 0 untuk menghindari pembagian 0
        abs_error = mp.fabs(actual[i] - predicted[i])
        percentage_error = abs_error / actual[i]
        sum_of_percentage_error += percentage_error
        MAPE = sum_of_percentage_error / (i + 1) * 100
        mape_list.append(float(MAPE))
    return mape_list

def fitting(closing_prices, stock_symbol):
    logging.debug(f'fitting called with closing_prices={closing_prices}, stock_symbol={stock_symbol}')
    Fitting_S_n_list = []
    v_list = []
    first_run = True
    
    # Check if we have enough data points
    if len(closing_prices) < 4:
        st.error("Tidak cukup data untuk melakukan fitting. Minimal 4 data point diperlukan.")
        return [], []
    
    for i in range(3):
        Fitting_S_n_list.append(float(closing_prices[i]))

    for i in range(3, len(closing_prices)):
        S_minus_1 = closing_prices[i - 3]
        S_0 = closing_prices[i - 2]
        S_1 = closing_prices[i - 1]
        S_2 = closing_prices[i]
        
        v_0 = determine_v_n(S_0, S_minus_1)
        v_1 = determine_v_n(S_1, S_0)
        v_2 = determine_v_n(S_2, S_1)
        
        if first_run:
            v_list.append(v_0)
            v_list.append(v_1)
            first_run = False
        v_list.append(v_2)

        try:
            alpha_n = determine_alpha_n(S_minus_1,S_0, S_1, S_2)
            beta_n = determine_beta_n(S_minus_1,S_1, S_2, alpha_n)
            h_n = determine_h_n(v_0, alpha_n, beta_n)
            condition_1 = (v_2 + (beta_n / alpha_n)) * v_2
            S_n = determine_s_n(S_minus_1, alpha_n, beta_n, h_n, condition_1, S_2, v_2, v_0)
        except (ZeroDivisionError) as e:
            logging.warning(f"Error in calculation at index {i}: {e}. Using fallback.")
            S_n = S_2  # fallback, data tidak berubah

        Fitting_S_n_list.append(float(S_n))
        logging.debug(f'Appended S_n={S_n} to Fitting_S_n_list')
    
    return Fitting_S_n_list, v_list

def forecasting(Fitting_S_n_list, start_date, end_date, stock_symbol):
    if len(Fitting_S_n_list) < 4:
        st.error("Tidak cukup data fitting untuk melakukan forecasting.")
        return [], []
        
    fitting_S_last = Fitting_S_n_list[-4:].copy()  # Make a copy to avoid modifying original
    
    try:
        closing_prices_full = get_data(stock_symbol, start_date, end_date)
        closing_prices_full = [price.item() for price in closing_prices_full]
        closing_prices_full = filter_prices_duplicates(closing_prices_full)
    except Exception as e:
        st.error(f"Error getting forecast data: {e}")
        return [], []
    
    forecast_days = len(closing_prices_full) - len(Fitting_S_n_list)
    
    if forecast_days <= 0:
        st.warning("Tidak terdapat cukup data untuk melakukan forecast.")
        return [], closing_prices_full[len(Fitting_S_n_list):]

    for i in range(3, forecast_days + 3):
        if i >= len(fitting_S_last):
            break
            
        S_minus_1 = fitting_S_last[i - 3]
        S_0 = fitting_S_last[i - 2]
        S_1 = fitting_S_last[i - 1]
        S_2 = fitting_S_last[i]
        
        v_0 = determine_v_n(S_0, S_minus_1)
        v_2 = determine_v_n(S_2, S_1)
        
        try:
            alpha_n = determine_alpha_n(S_minus_1,S_0, S_1, S_2)
            beta_n = determine_beta_n(S_0, S_1, S_2, alpha_n)
            h_n = determine_h_n(v_0, alpha_n, beta_n)
            condition_1 = (v_2 + (beta_n / alpha_n)) * v_2
            S_n = determine_s_n(S_minus_1, alpha_n, beta_n, h_n, condition_1, S_2, v_2, v_0)
        except (ZeroDivisionError) as e:
            logging.warning(f"Error in forecast at step {i}: {e}. Using previous value.")
            S_n = S_2
            
        fitting_S_last.append(float(S_n))
    
    forecast_S_list = fitting_S_last[3:] 
    closing_forecast = closing_prices_full[len(Fitting_S_n_list)-1:] # mau hilangin (-1) tapi error di length harus tanya (!!!)
    return forecast_S_list, closing_forecast

def filter_prices_duplicates(closing_prices):
    if not closing_prices:
        return []
    
    filtered_prices = [closing_prices[0]]
    for i in range(1, len(closing_prices)):
        if closing_prices[i] != closing_prices[i-1]:
            filtered_prices.append(closing_prices[i])
    return filtered_prices

def create_excel_download(stock_symbol, start_date, end_date, forecast_end_date, 
                         closing_prices, Fitting_S_n_list, S_forecast, closing_forecast):
    """
    Create Excel file with actual dates from Yahoo Finance data, fitting data, and forecast data
    """
    try:
        # Get actual dates from Yahoo Finance for fitting period
        fitting_data = yf.download(stock_symbol, start=start_date, end=end_date, auto_adjust=False)
        fitting_dates = fitting_data.index.tolist()
        
        # Filter dates to match the length of closing_prices (after duplicate removal)
        if len(fitting_dates) > len(closing_prices):
            # If we have more dates than prices (due to duplicate filtering), 
            # we need to align them properly
            fitting_dates = fitting_dates[:len(closing_prices)]
        
        # Create DataFrame for fitting period
        fitting_df = pd.DataFrame({
            'Date': fitting_dates,
            'Actual_Price': closing_prices,
            'Fitted_Price': Fitting_S_n_list[:len(closing_prices)] if len(Fitting_S_n_list) >= len(closing_prices) else Fitting_S_n_list + [None] * (len(closing_prices) - len(Fitting_S_n_list)),
            'Type': 'Fitting'
        })
        
        # Create DataFrame for forecast period if forecast data exists
        if S_forecast and closing_forecast:
            # Get actual dates from Yahoo Finance for forecast period
            forecast_data = yf.download(stock_symbol, start=end_date, end=forecast_end_date, auto_adjust=False)
            forecast_dates = forecast_data.index.tolist()
            
            # Remove the first date if it overlaps with fitting period
            if forecast_dates and fitting_dates and forecast_dates[0] <= fitting_dates[-1]:
                forecast_dates = forecast_dates[1:]
            
            # Align forecast dates with forecast data length
            min_forecast_len = min(len(forecast_dates), len(closing_forecast), len(S_forecast))
            forecast_dates = forecast_dates[:min_forecast_len]
            closing_forecast = closing_forecast[:min_forecast_len]
            S_forecast = S_forecast[:min_forecast_len]
            
            forecast_df = pd.DataFrame({
                'Date': forecast_dates,
                'Actual_Price': closing_forecast,
                'Fitted_Price': None,
                'Forecast_Price': S_forecast,
                'Type': 'Forecast'
            })
            
            # Add Forecast_Price column to fitting_df for consistency
            fitting_df['Forecast_Price'] = None
            
            # Combine both DataFrames
            combined_df = pd.concat([fitting_df, forecast_df], ignore_index=True)
        else:
            # If no forecast data, just use fitting data
            fitting_df['Forecast_Price'] = None
            combined_df = fitting_df
        
        # Reorder columns
        combined_df = combined_df[['Date', 'Actual_Price', 'Fitted_Price', 'Forecast_Price', 'Type']]
        
        # Create Excel file in memory
        output = BytesIO()
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{stock_symbol}_Analysis"
        
        # Add title and metadata
        ws['A1'] = f"Stock Analysis Report - {stock_symbol}"
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Fitting Period: {start_date} to {end_date}"
        ws['A4'] = f"Forecast Period: {end_date} to {forecast_end_date}"
        ws['A5'] = ""  # Empty row
        
        # Style the title
        title_font = Font(size=14, bold=True)
        ws['A1'].font = title_font
        
        # Add headers starting from row 6
        headers = ['Date', 'Actual Price', 'Fitted Price', 'Forecast Price', 'Type']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data starting from row 7
        for row_idx, (_, row) in enumerate(combined_df.iterrows(), 7):
            # Format date properly
            if pd.notna(row['Date']):
                if hasattr(row['Date'], 'strftime'):
                    ws.cell(row=row_idx, column=1, value=row['Date'].strftime('%Y-%m-%d'))
                else:
                    ws.cell(row=row_idx, column=1, value=str(row['Date']))
            else:
                ws.cell(row=row_idx, column=1, value="")
                
            ws.cell(row=row_idx, column=2, value=row['Actual_Price'])
            ws.cell(row=row_idx, column=3, value=row['Fitted_Price'])
            ws.cell(row=row_idx, column=4, value=row['Forecast_Price'])
            ws.cell(row=row_idx, column=5, value=row['Type'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        logging.error(f"Error in create_excel_download: {e}")

#--------------------------------------------------------------------------------------------------------------#

# Streamlit UI - Main Page Layout (Simplified Design)
st.title("📈 Stock Price Fitting and Forecasting Web")

st.markdown("---")

# Input Parameter
st.subheader("📋 Input Parameter")

# Create 4 columns for the input fields
col1, col2, col3, col4 = st.columns(4)

with col1:
    st.markdown("**Stock Symbol**")
    stock_symbol = st.text_input("", value="BBCA.JK", key="stock_input", label_visibility="collapsed")

with col2:
    st.markdown("**Fitting Start Date**")
    start_date = st.date_input("", value=datetime(2024, 1, 1), key="start_date_input", label_visibility="collapsed")

with col3:
    st.markdown("**Fitting Period (Day)**")
    training_days = st.number_input("", min_value=1, max_value=365, value=120, step=1, key="training_days", label_visibility="collapsed")

with col4:
    st.markdown("**Forecast Period (Day)**")
    forecast_days = st.number_input("", min_value=1, max_value=365, value=60, step=1, key="forecast_days", label_visibility="collapsed")

# Calculate dates based on training and forecast days
end_date = start_date + timedelta(days=training_days)
forecast_end_date = end_date + timedelta(days=forecast_days)

# Advanced Options (Collapsible)
with st.expander("⚙️ Advanced Options"):
    col_adv1, col_adv2 = st.columns(2)
    
    with col_adv1:
        use_custom_end = st.checkbox("Custom Fitting End Date", value=False)
        if use_custom_end:
            st.markdown("**Custom Fitting End Date**")
            custom_end_date = st.date_input("", value=datetime(2024, 4, 30), key="custom_end", label_visibility="collapsed")
            end_date = custom_end_date
    
    with col_adv2:
        use_custom_forecast_end = st.checkbox("Custom Forecast End Date", value=False)
        if use_custom_forecast_end:
            st.markdown("**Custom Forecast End Date**")
            custom_forecast_end = st.date_input("", value=datetime(2024, 6, 29), key="custom_forecast_end", label_visibility="collapsed")
            forecast_end_date = custom_forecast_end

# Ringkasan Input
col_summary1, col_summary2 = st.columns(2)

with col_summary1:
    st.markdown(f"""
    **📊 Fitting Period Details:**
    - **Periode Fitting:** {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}
    - **Durasi Fitting:** {training_days} hari
    """)

with col_summary2:
    st.markdown(f"""
    **🔮 Data Forecast Period Details:**
    - **Periode Forecast:** {end_date.strftime('%d/%m/%Y')} - {forecast_end_date.strftime('%d/%m/%Y')}
    - **Durasi Forecast:** {forecast_days} hari
    """)

st.markdown("---")


# Validation
if start_date >= end_date:
    st.error("Start date harus lebih kecil dari end date!")
elif end_date >= forecast_end_date:
    st.error("End date harus lebih kecil dari forecast end date!")

# Run Analysis Button (Full Width, Red/Primary Color)
run_forecast = st.button("🔗 Submit Data", use_container_width=True, type="primary")

# Rest of the code remains the same for the analysis execution...
if run_forecast:
    try:
        with st.spinner("Mengambil dan memproses data..."):
            # Get initial data
            closing_prices = get_data(stock_symbol, start_date, end_date)
            closing_prices = [price.item() for price in closing_prices]
            closing_prices = filter_prices_duplicates(closing_prices)
            
            if len(closing_prices) < 4:
                st.error("Data tidak cukup untuk melakukan forecasting. Minimal 4 data point diperlukan.")
                st.stop()

            # FITTING
            Fitting_S_n_list, v_list = fitting(closing_prices, stock_symbol)
            
            if not Fitting_S_n_list:
                st.error("Gagal melakukan fitting data.")
                st.stop()
                
            mape_fit = determine_MAPE_list(closing_prices, Fitting_S_n_list)

            # FORECASTING
            S_forecast, closing_forecast = forecasting(
                Fitting_S_n_list,
                start_date.strftime("%Y-%m-%d"),
                forecast_end_date.strftime("%Y-%m-%d"),
                stock_symbol
            )
            
            if S_forecast and closing_forecast:
                mape_forecast = determine_MAPE_list(closing_forecast, S_forecast)
            else:
                mape_forecast = []

        st.success("Selesai!")

        # Display results only if we have valid data
        if Fitting_S_n_list:

            # Display summary statistics
            st.subheader("📊 Statistic Details")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("Jumlah Data Fitting", len(closing_prices))
            with col2:
                if mape_fit:
                    st.metric("MAPE Fitting", f"{np.mean(mape_fit):.2f}%")
            with col3:
                if mape_forecast:
                    st.metric("MAPE Forecast", f"{np.mean(mape_forecast):.2f}%")
            with col4:
                st.metric("Periode Forecast", f"{(forecast_end_date - end_date).days} hari")

            # Grafik 1: Fitting vs Actual
            st.subheader(f"📊 Grafik Fitting vs Actual ({stock_symbol})")
            fig_fit, ax_fit = plt.subplots(figsize=(10, 6))
            ax_fit.plot(closing_prices, label="Actual", color='black', linewidth=2)
            ax_fit.plot(Fitting_S_n_list, label="Fitted", color='blue', linewidth=2)
            ax_fit.set_title(f"Fitting Data Harga Saham ({stock_symbol})")
            ax_fit.set_xlabel("Hari")
            ax_fit.set_ylabel("Harga")
            ax_fit.legend()
            ax_fit.grid(True, alpha=0.3)
            st.pyplot(fig_fit)

            # Grafik 2: Fitting + Forecast vs Actual (only if forecast data exists)
            if S_forecast and closing_forecast:
                st.subheader(f"📈 Grafik Fitting + Forecast vs Actual ({stock_symbol})")
                fig_forecast, ax_forecast = plt.subplots(figsize=(12, 6))
                
                # Plot actual data
                all_actual = closing_prices + closing_forecast
                ax_forecast.plot(all_actual, label="Actual", color='black', linewidth=2)
                
                # Plot fitted data
                ax_forecast.plot(range(len(Fitting_S_n_list)), Fitting_S_n_list, label="Fitted", color='blue', linewidth=2)
                
                # Plot forecast data
                forecast_start_idx = len(Fitting_S_n_list)
                forecast_end_idx = forecast_start_idx + len(S_forecast)
                ax_forecast.plot(range(forecast_start_idx, forecast_end_idx), S_forecast, 
                               label="Forecast", color='orange', linewidth=2)
                
                # Add vertical line to show where forecast starts 
                ax_forecast.axvline(x=len(closing_prices), color='red', linestyle='--', 
                                  label='Forecast Start', alpha=0.7)
                
                ax_forecast.set_title(f"Fitting dan Forecast Harga Saham ({stock_symbol})")
                ax_forecast.set_xlabel("Hari")
                ax_forecast.set_ylabel("Harga")
                ax_forecast.legend()
                ax_forecast.grid(True, alpha=0.3)
                st.pyplot(fig_forecast)

            # MAPE Fitting
            if mape_fit:
                st.subheader(f"📉 Hasil MAPE Fitting - Rata-rata: {np.mean(mape_fit):.2f}%")
                fig_mape_fit, ax_mape_fit = plt.subplots(figsize=(10, 6))
                ax_mape_fit.plot(mape_fit, color='purple', label='MAPE Fitting (%)', linewidth=2)
                ax_mape_fit.set_title(f"Grafik MAPE Selama Fitting ({stock_symbol})")
                ax_mape_fit.set_xlabel("Hari")
                ax_mape_fit.set_ylabel("MAPE (%)")
                ax_mape_fit.legend()
                ax_mape_fit.grid(True, alpha=0.3)
                st.pyplot(fig_mape_fit)

            # MAPE Forecast
            if mape_forecast:
                st.subheader(f"📉 Hasil MAPE Forecast - Rata-rata: {np.mean(mape_forecast):.2f}%")
                fig_mape_forecast, ax_mape_forecast = plt.subplots(figsize=(10, 6))
                ax_mape_forecast.plot(mape_forecast, color='orange', label='MAPE Forecast (%)', linewidth=2)
                ax_mape_forecast.set_title(f"Grafik MAPE Selama Forecasting ({stock_symbol})")
                ax_mape_forecast.set_xlabel("Hari")
                ax_mape_forecast.set_ylabel("MAPE (%)")
                ax_mape_forecast.legend()
                ax_mape_forecast.grid(True, alpha=0.3)
                st.pyplot(fig_mape_forecast)

            # Create download button for Excel
            st.subheader("💾 Download Data")
                    
            try:
                excel_data = create_excel_download(
                    stock_symbol=stock_symbol,
                    start_date=start_date,
                    end_date=end_date,
                    forecast_end_date=forecast_end_date,
                    closing_prices=closing_prices,
                    Fitting_S_n_list=Fitting_S_n_list,
                    S_forecast=S_forecast if S_forecast else [],
                    closing_forecast=closing_forecast if closing_forecast else []
                )
                        
                filename = f"{stock_symbol}_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        
                st.download_button(
                    label="📥 Download Excel Report",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Download complete analysis data in Excel format"
                )
                        
                st.info(f"📊 File akan berisi data dari {start_date} hingga {forecast_end_date}")
                        
            except Exception as e:
                st.error(f"Error creating Excel file: {str(e)}")
                logging.error(f"Excel creation error: {e}")
                
    except Exception as e:
        st.error(f"Terjadi kesalahan: {str(e)}")
        logging.error(f"Main execution error: {e}")
        st.info("Silakan coba dengan parameter yang berbeda atau periksa koneksi data.")